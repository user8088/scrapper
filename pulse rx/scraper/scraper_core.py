"""
High-end product description scraper for Pulse Rx.

Matches each product folder (named by Item Id) against two pharmacy sites and
fetches a rich product description:

  * dwatson.pk   -> Magento store. Search at /catalogsearch/result/?q=...
                    The description is captured as HTML from the page body
                    block (preserving <b> headings and <br> line structure);
                    the plain-text `og:description` meta is only a fallback.
  * healthwire.pk-> Rails store. JSON search at /searches. Its `item_id`
                    matches the Pulse Rx Item Id exactly, giving an
                    authoritative match. Description is built as HTML from the
                    product page body sections (Description / Uses / ...),
                    preserving paragraphs and bullet lists exactly.

Descriptions are written as sanitized HTML fragments restricted to the tag
subset the Pulse Rx site renders verbatim (see `clean_html`), so the storefront
shows an exact copy of the source site's structure.

The module is UI-agnostic: `process_product` returns a structured result and
`write_description` commits a description into the product's xlsx file.
"""

from __future__ import annotations

import csv
import os
import re
import time
import json
import random
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from typing import Callable, Optional

from urllib.parse import urljoin

import requests
import urllib3
from bs4 import BeautifulSoup, Comment
from rapidfuzz import fuzz
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

# Sources
SOURCE_DWATSON = "dwatson"
SOURCE_HEALTHWIRE = "healthwire"

# Statuses
STATUS_FOUND = "found"
STATUS_NOT_FOUND = "not_found"
STATUS_NO_CSV = "no_csv_row"
STATUS_ERROR = "error"
STATUS_LOW_CONFIDENCE = "low_confidence"


# --------------------------------------------------------------------------- #
# Text helpers
# --------------------------------------------------------------------------- #
_STRENGTH_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(mcg|mg|ml|gm|g|iu|%|mg/ml)", re.I)
_PACK_RE = re.compile(r"(\d+)\s*(?:s\b|tablets?|capsules?|caps?|tabs?|sachets?|pcs?|pieces?|ampoules?|vials?)", re.I)


def normalize_text(s: str) -> str:
    if not s:
        return ""
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def strength_tokens(name: str) -> set:
    """Extract normalized strength tokens like {'18mcg', '500mg', '60ml'}."""
    out = set()
    for num, unit in _STRENGTH_RE.findall(name or ""):
        num = num.rstrip("0").rstrip(".") if "." in num else num
        out.add(f"{num}{unit.lower()}")
    return out


def pack_tokens(name: str) -> set:
    return {m for m in _PACK_RE.findall(name or "")}


def name_score(a: str, b: str) -> float:
    return fuzz.token_sort_ratio((a or "").lower(), (b or "").lower())


# --------------------------------------------------------------------------- #
# HTML helpers
# --------------------------------------------------------------------------- #
# Tag subset the Pulse Rx site renders verbatim (mirrors the DOMPurify
# allowlist in lib/products/description-content.ts). Anything written within
# this subset reaches the storefront unchanged.
_HTML_KEEP_TAGS = {
    "p", "br", "strong", "b", "em", "i", "u",
    "ul", "ol", "li", "h1", "h2", "h3", "h4", "a",
}
# Non-content tags whose subtree should be removed entirely; every other
# disallowed tag is unwrapped so its text/children survive.
_HTML_DROP_TAGS = {
    "script", "style", "noscript", "iframe", "img", "svg", "picture",
    "source", "video", "audio", "form", "button", "input", "select",
}


def clean_html(node, base_url: str = "") -> str:
    """Reduce an HTML fragment to the allowlisted tag subset, preserving the
    source structure exactly (paragraphs, bullet lists, bold headings, line
    breaks). All attributes are stripped except a[href], which is resolved
    against base_url so relative links keep working off-site."""
    soup = BeautifulSoup(str(node), "html.parser")
    for comment in soup.find_all(string=lambda s: isinstance(s, Comment)):
        comment.extract()
    for tag in soup.find_all(list(_HTML_DROP_TAGS)):
        tag.decompose()
    for tag in soup.find_all(True):
        if tag.name not in _HTML_KEEP_TAGS:
            tag.unwrap()
            continue
        if tag.name == "a":
            href = (tag.get("href") or "").strip()
            if href and base_url and not href.startswith(("http://", "https://", "mailto:", "tel:")):
                href = urljoin(base_url, href)
            tag.attrs = {"href": href} if href else {}
        else:
            tag.attrs = {}
    html = str(soup)
    # Collapse whitespace runs: rendering ignores them and they bloat the cell.
    return re.sub(r"\s+", " ", html).strip()


def html_text(html: str) -> str:
    """Plain-text view of an HTML fragment (previews, emptiness checks)."""
    if not html:
        return ""
    if "<" not in html:
        return html.strip()
    return BeautifulSoup(html, "html.parser").get_text(" ", strip=True)


def is_strong_match(query_name: str, candidate_name: str, min_score: float) -> tuple[bool, float, str]:
    """
    Decide whether candidate_name is the same product as query_name.
    Requires a good fuzzy score AND consistent strength tokens (no conflicts).
    Returns (ok, score, reason).
    """
    score = name_score(query_name, candidate_name)
    q_str = strength_tokens(query_name)
    c_str = strength_tokens(candidate_name)

    # If both have strength tokens, they must overlap (no conflicting dosage).
    if q_str and c_str and not (q_str & c_str):
        return False, score, f"strength mismatch {sorted(q_str)} vs {sorted(c_str)}"

    if score < min_score:
        return False, score, f"score {score:.0f} < {min_score:.0f}"

    return True, score, "ok"


# --------------------------------------------------------------------------- #
# Catalog (CSV)
# --------------------------------------------------------------------------- #
@dataclass
class CatalogItem:
    item_id: str
    name: str
    generic: str = ""
    manufacturer: str = ""
    category: str = ""


def load_catalog(csv_path: str) -> dict[str, CatalogItem]:
    catalog: dict[str, CatalogItem] = {}
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            iid = (row.get("Item Id") or "").strip()
            if not iid:
                continue
            catalog[iid] = CatalogItem(
                item_id=iid,
                name=(row.get("Item Name") or "").strip(),
                generic=(row.get("Genaric Name") or row.get("Generic Name") or "").strip(),
                manufacturer=(row.get("Manufacturer") or "").strip(),
                category=(row.get("B2B Category") or row.get("Category") or "").strip(),
            )
    return catalog


def list_product_folders(products_root: str) -> list[str]:
    out = []
    for d in os.listdir(products_root):
        if os.path.isdir(os.path.join(products_root, d)) and d.isdigit():
            out.append(d)
    out.sort(key=lambda x: int(x))
    return out


# --------------------------------------------------------------------------- #
# Site clients
# --------------------------------------------------------------------------- #
class DwatsonClient:
    SEARCH_URL = "https://dwatson.pk/catalogsearch/result/"

    def __init__(self, session: requests.Session, timeout: int = 40):
        self.s = session
        self.timeout = timeout

    def search(self, query: str) -> list[tuple[str, str]]:
        r = self.s.get(self.SEARCH_URL, params={"q": query}, timeout=self.timeout, verify=False)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")
        out = []
        for a in soup.select("a.product-item-link"):
            title = a.get_text(strip=True)
            href = a.get("href")
            if title and href:
                out.append((title, href))
        return out

    def fetch_description(self, url: str) -> str:
        r = self.s.get(url, timeout=self.timeout, verify=False)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")
        # The body block carries the real formatting (<b> headings, <br> line
        # structure); og:description is Magento's plain-text flattening of the
        # same content, so it is only a fallback.
        node = soup.select_one(".product.attribute.description .value, .description .value, #description")
        if node:
            html = clean_html(node, base_url=url)
            if html_text(html):
                return html
        og = soup.find("meta", attrs={"property": "og:description"})
        if og and og.get("content"):
            return normalize_text(og["content"])
        return ""


class HealthwireClient:
    SEARCH_URL = "https://healthwire.pk/searches"
    BASE = "https://healthwire.pk"
    ATTRS = ["id", "name", "chemical_name", "image", "category", "actual_price",
             "discounted_price", "url", "manufacturer_name", "item_variant_id", "item_id"]
    # Sections to assemble into the description, in display order.
    SECTIONS = ["Description", "Uses", "How To Use", "Dosage", "Side Effects",
                "Precautions & Warnings", "Drug Interactions", "Storage/Disposal"]

    def __init__(self, session: requests.Session, timeout: int = 40):
        self.s = session
        self.timeout = timeout

    def search(self, query: str) -> list[dict]:
        params = [("searches[0][q]", query), ("searches[0][model]", "items")]
        for a in self.ATTRS:
            params.append(("searches[0][attributes][]", a))
        r = self.s.get(self.SEARCH_URL, params=params, timeout=self.timeout, verify=False,
                       headers={"X-Requested-With": "XMLHttpRequest",
                                "Accept": "application/json, text/javascript, */*; q=0.01"})
        r.raise_for_status()
        data = r.json()
        if isinstance(data, list) and data:
            return [h.get("_source", {}) for h in data[0].get("hits", [])]
        return []

    def fetch_description(self, url: str) -> str:
        full = url if url.startswith("http") else self.BASE + url
        r = self.s.get(full, timeout=self.timeout, verify=False)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")
        parts: list[str] = []
        for h2 in soup.find_all("h2"):
            title = h2.get_text(strip=True)
            if title not in self.SECTIONS:
                continue
            chunk: list[str] = []
            sib = h2.find_next_sibling()
            guard = 0
            while sib is not None and sib.name not in ("h2",) and guard < 12:
                html = clean_html(sib, base_url=self.BASE)
                if html:
                    chunk.append(html)
                sib = sib.find_next_sibling()
                guard += 1
            body = "".join(chunk)
            text = html_text(body)
            if text and text.lower() not in ("n/a", "na", "-", "not available"):
                parts.append(f"<h2>{title}</h2>{body}")
        return "".join(parts)


# --------------------------------------------------------------------------- #
# Result + config
# --------------------------------------------------------------------------- #
@dataclass
class MatchResult:
    folder_id: str
    csv_name: str = ""
    status: str = STATUS_NOT_FOUND
    source: str = ""
    matched_name: str = ""
    score: float = 0.0
    url: str = ""
    description: str = ""
    note: str = ""
    selected: bool = True  # for UI preview / commit

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class ScrapeConfig:
    products_root: str
    csv_path: str
    header: str = "Details"               # xlsx column header to keep
    batch_size: int = 50
    source_order: tuple = (SOURCE_DWATSON, SOURCE_HEALTHWIRE)
    min_score: float = 80.0               # dwatson strict match threshold
    request_delay: float = 0.8            # polite delay before each product (s)
    timeout: int = 40
    skip_filled: bool = False             # skip products that already have content
    overwrite: bool = True
    max_workers: int = 4                  # concurrent products (1 = sequential)


# --------------------------------------------------------------------------- #
# Engine
# --------------------------------------------------------------------------- #
class ScraperEngine:
    def __init__(self, config: ScrapeConfig):
        self.cfg = config
        self.catalog = load_catalog(config.csv_path)
        # Each worker thread gets its own Session + clients. requests.Session is
        # not guaranteed thread-safe, and the read-only catalog is shared freely.
        self._local = threading.local()

    def _clients(self) -> tuple["DwatsonClient", "HealthwireClient"]:
        if not getattr(self._local, "ready", False):
            session = requests.Session()
            session.headers.update({"User-Agent": USER_AGENT})
            self._local.dwatson = DwatsonClient(session, self.cfg.timeout)
            self._local.healthwire = HealthwireClient(session, self.cfg.timeout)
            self._local.ready = True
        return self._local.dwatson, self._local.healthwire

    # -- per-source resolution ------------------------------------------------ #
    def _resolve_dwatson(self, item: CatalogItem, client: "DwatsonClient") -> Optional[MatchResult]:
        results = client.search(item.name)
        if not results:
            return None
        best = None
        best_score = -1.0
        for title, href in results:
            ok, score, reason = is_strong_match(item.name, title, self.cfg.min_score)
            if ok and score > best_score:
                best, best_score = (title, href), score
        if not best:
            return None
        desc = client.fetch_description(best[1])
        if not desc:
            return None
        return MatchResult(
            folder_id=item.item_id, csv_name=item.name, status=STATUS_FOUND,
            source=SOURCE_DWATSON, matched_name=best[0], score=round(best_score, 1),
            url=best[1], description=desc,
        )

    def _resolve_healthwire(self, item: CatalogItem, client: "HealthwireClient") -> Optional[MatchResult]:
        hits = client.search(item.name)
        if not hits:
            return None
        # Authoritative: match by item_id == folder id.
        chosen = next((h for h in hits if str(h.get("item_id")) == item.item_id), None)
        score = 100.0
        if chosen is None:
            # Fall back to best fuzzy hit with strength validation.
            best = None
            best_score = -1.0
            for h in hits:
                ok, sc, _ = is_strong_match(item.name, h.get("name", ""), self.cfg.min_score)
                if ok and sc > best_score:
                    best, best_score = h, sc
            if best is None:
                return None
            chosen, score = best, round(best_score, 1)
        url = chosen.get("url") or ""
        if not url:
            return None
        desc = client.fetch_description(url)
        if not desc:
            return None
        return MatchResult(
            folder_id=item.item_id, csv_name=item.name, status=STATUS_FOUND,
            source=SOURCE_HEALTHWIRE, matched_name=chosen.get("name", ""), score=score,
            url=client.BASE + url if url.startswith("/") else url, description=desc,
        )

    def process_product(self, folder_id: str) -> MatchResult:
        item = self.catalog.get(folder_id)
        if not item or not item.name:
            return MatchResult(folder_id=folder_id, status=STATUS_NO_CSV,
                               note="No matching row in CSV")
        dwatson, healthwire = self._clients()
        resolvers = {
            SOURCE_DWATSON: (self._resolve_dwatson, dwatson),
            SOURCE_HEALTHWIRE: (self._resolve_healthwire, healthwire),
        }
        last_error = ""
        for src in self.cfg.source_order:
            entry = resolvers.get(src)
            if not entry:
                continue
            fn, client = entry
            try:
                res = fn(item, client)
                if res and res.description:
                    return res
            except Exception as exc:  # network / parse error -> try next source
                last_error = f"{src}: {exc}"
        return MatchResult(folder_id=folder_id, csv_name=item.name,
                           status=STATUS_ERROR if last_error else STATUS_NOT_FOUND,
                           note=last_error or "No match on either site")

    # -- batch ---------------------------------------------------------------- #
    def select_folders(self, offset: int = 0, limit: Optional[int] = None) -> list[str]:
        folders = list_product_folders(self.cfg.products_root)
        if self.cfg.skip_filled:
            folders = [f for f in folders if not _xlsx_has_content(self.cfg.products_root, f)]
        limit = limit if limit is not None else self.cfg.batch_size
        return folders[offset: offset + limit]

    def _process_with_delay(self, fid: str, stop: Callable[[], bool]) -> Optional[MatchResult]:
        if stop():
            return None
        # Polite, jittered pause before each product so concurrent workers don't
        # fire in lockstep. Aggregate request rate ~= workers / request_delay.
        if self.cfg.request_delay:
            time.sleep(random.uniform(0, self.cfg.request_delay))
        return self.process_product(fid)

    def run_batch(
        self,
        folder_ids: list[str],
        on_result: Optional[Callable[[MatchResult], None]] = None,
        should_stop: Optional[Callable[[], bool]] = None,
    ) -> list[MatchResult]:
        out: list[MatchResult] = []
        stop = should_stop or (lambda: False)
        workers = max(1, int(getattr(self.cfg, "max_workers", 1)))

        # Sequential path preserves the original, gentlest behaviour.
        if workers == 1:
            for fid in folder_ids:
                if stop():
                    break
                res = self.process_product(fid)
                out.append(res)
                if on_result:
                    on_result(res)
                if self.cfg.request_delay:
                    time.sleep(self.cfg.request_delay)
            return out

        # Concurrent path: N products in flight at once. as_completed yields on
        # the calling thread, so on_result/out are touched single-threaded here;
        # only process_product runs in the pool (each worker has its own session).
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futures = [ex.submit(self._process_with_delay, fid, stop) for fid in folder_ids]
            for fut in as_completed(futures):
                res = fut.result()
                if res is None:  # skipped because stop was requested
                    continue
                out.append(res)
                if on_result:
                    on_result(res)
        return out


# --------------------------------------------------------------------------- #
# xlsx writing
# --------------------------------------------------------------------------- #
def _xlsx_path(products_root: str, folder_id: str) -> str:
    return os.path.join(products_root, folder_id, "descriptions", "product_details.xlsx")


def _xlsx_has_content(products_root: str, folder_id: str) -> bool:
    path = _xlsx_path(products_root, folder_id)
    if not os.path.exists(path):
        return False
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        val = ws.cell(2, 1).value
        wb.close()
        return bool(val and str(val).strip())
    except Exception:
        return False


XLSX_CELL_LIMIT = 32767  # Excel's hard cap on characters per cell


def write_description(products_root: str, folder_id: str, description: str,
                      header: str = "Details") -> str:
    """Write the description into A2 (header in A1). Returns the file path."""
    if len(description) > XLSX_CELL_LIMIT:
        # Cut at a tag boundary; the site's sanitizer rebalances unclosed tags.
        cut = description.rfind("</", 0, XLSX_CELL_LIMIT)
        description = description[:cut] if cut > 0 else description[:XLSX_CELL_LIMIT]
    path = _xlsx_path(products_root, folder_id)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if os.path.exists(path):
        try:
            wb = load_workbook(path)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()
    ws = wb.active
    existing_header = ws.cell(1, 1).value
    ws.cell(row=1, column=1, value=existing_header if existing_header else header)
    cell = ws.cell(row=2, column=1, value=description)
    cell.alignment = cell.alignment.copy(wrap_text=True, vertical="top")
    wb.save(path)
    return path


def read_existing_description(products_root: str, folder_id: str) -> str:
    path = _xlsx_path(products_root, folder_id)
    if not os.path.exists(path):
        return ""
    try:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        val = ws.cell(2, 1).value
        wb.close()
        return str(val) if val else ""
    except Exception:
        return ""


# --------------------------------------------------------------------------- #
# Report export
# --------------------------------------------------------------------------- #
STATUS_LABELS = {
    STATUS_FOUND: "Found",
    STATUS_NOT_FOUND: "Not found",
    STATUS_NO_CSV: "No CSV row",
    STATUS_ERROR: "Error",
    STATUS_LOW_CONFIDENCE: "Low confidence",
}

# (cell fill, font colour) per status -- Excel's familiar green/amber/red palette.
_REPORT_STYLE = {
    STATUS_FOUND: ("C6EFCE", "1B7A30"),
    STATUS_NOT_FOUND: ("FFEB9C", "9C6500"),
    STATUS_ERROR: ("FFC7CE", "9C0006"),
    STATUS_NO_CSV: ("E2E2E2", "5A5A5A"),
    STATUS_LOW_CONFIDENCE: ("FFEB9C", "9C6500"),
}

_HEADER_FILL = "305496"     # dark blue
_TITLE_FILL = "1F3864"      # darker blue
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def compute_stats(results) -> dict:
    """Aggregate counts/rates from a list of MatchResult."""
    total = len(results)
    by_status: dict[str, int] = {}
    by_source: dict[str, int] = {}
    for r in results:
        by_status[r.status] = by_status.get(r.status, 0) + 1
        if r.status == STATUS_FOUND and r.source:
            by_source[r.source] = by_source.get(r.source, 0) + 1
    found = by_status.get(STATUS_FOUND, 0)
    return {
        "total": total,
        "found": found,
        "not_found": by_status.get(STATUS_NOT_FOUND, 0),
        "errors": by_status.get(STATUS_ERROR, 0),
        "no_csv": by_status.get(STATUS_NO_CSV, 0),
        "low_confidence": by_status.get(STATUS_LOW_CONFIDENCE, 0),
        "by_source": by_source,
        "success_rate": (found / total * 100.0) if total else 0.0,
    }


def export_report(results, path: str, meta: Optional[dict] = None) -> str:
    """Write a polished two-sheet xlsx report (Summary + Results) and return path.

    Results are colour-marked by status, source URLs are clickable hyperlinks,
    and the Summary sheet carries the run statistics.
    """
    meta = meta or {}
    rows = sorted(
        results,
        key=lambda r: (int(r.folder_id) if str(r.folder_id).isdigit() else 1 << 62,
                       str(r.folder_id)),
    )
    stats = compute_stats(results)

    wb = Workbook()
    _build_summary_sheet(wb.active, stats, meta)
    _build_results_sheet(wb.create_sheet("Results"), rows)
    wb.save(path)
    return path


def _build_summary_sheet(ws, stats: dict, meta: dict) -> None:
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 46

    title = ws.cell(1, 1, "Pulse Rx — Scrape Report")
    title.font = Font(bold=True, size=16, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor=_TITLE_FILL)
    title.alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells("A1:B1")
    ws.cell(1, 2).fill = PatternFill("solid", fgColor=_TITLE_FILL)
    ws.row_dimensions[1].height = 28

    r = 3
    info = [
        ("Generated", meta.get("generated", "")),
        ("App version", meta.get("app_version", "")),
        ("Products folder", meta.get("products_root", "")),
        ("Source priority", meta.get("source_order", "")),
        ("Threads", meta.get("threads", "")),
    ]
    for label, value in info:
        if value in ("", None):
            continue
        ws.cell(r, 1, label).font = Font(color="808080")
        ws.cell(r, 2, str(value))
        r += 1

    r += 1
    hdr = ws.cell(r, 1, "Statistics")
    hdr.font = Font(bold=True, size=12, color="FFFFFF")
    hdr.fill = PatternFill("solid", fgColor=_HEADER_FILL)
    ws.cell(r, 2).fill = PatternFill("solid", fgColor=_HEADER_FILL)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1

    metrics = [
        ("Total products", stats["total"], None),
        ("Found", stats["found"], STATUS_FOUND),
        ("Not found", stats["not_found"], STATUS_NOT_FOUND),
        ("Errors", stats["errors"], STATUS_ERROR),
        ("No CSV row", stats["no_csv"], STATUS_NO_CSV),
    ]
    if stats["low_confidence"]:
        metrics.append(("Low confidence", stats["low_confidence"], STATUS_LOW_CONFIDENCE))
    metrics.append(("Success rate", f"{stats['success_rate']:.1f}%", None))

    for label, value, status in metrics:
        c_label = ws.cell(r, 1, label)
        c_value = ws.cell(r, 2, value)
        c_label.border = _BORDER
        c_value.border = _BORDER
        c_label.font = Font(bold=(label in ("Total products", "Success rate")))
        if status and status in _REPORT_STYLE:
            fill, font = _REPORT_STYLE[status]
            for c in (c_label, c_value):
                c.fill = PatternFill("solid", fgColor=fill)
                c.font = Font(color=font, bold=c is c_value)
        r += 1

    if stats["by_source"]:
        r += 1
        sub = ws.cell(r, 1, "Descriptions by source")
        sub.font = Font(bold=True, color="FFFFFF")
        sub.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        ws.cell(r, 2).fill = PatternFill("solid", fgColor=_HEADER_FILL)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        r += 1
        for src, n in sorted(stats["by_source"].items(), key=lambda kv: -kv[1]):
            ws.cell(r, 1, src).border = _BORDER
            ws.cell(r, 2, n).border = _BORDER
            r += 1


def _build_results_sheet(ws, rows) -> None:
    ws.sheet_view.showGridLines = False
    headers = ["Item Id", "CSV Name", "Status", "Source", "Matched Name",
               "Score", "Source URL", "Description"]
    widths = [10, 34, 13, 12, 34, 8, 52, 90]
    for col, (text, width) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(1, col, text)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        c.alignment = Alignment(vertical="center", horizontal="center")
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for i, res in enumerate(rows, start=2):
        values = [
            res.folder_id,
            res.csv_name,
            STATUS_LABELS.get(res.status, res.status),
            res.source or "",
            res.matched_name or "",
            (round(res.score, 1) if res.score else ""),
            res.url or "",
            res.description or res.note or "",
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(i, col, val)
            c.border = _BORDER
            c.alignment = Alignment(
                vertical="top",
                wrap_text=(col in (2, 5, 8)),
                horizontal=("center" if col in (3, 4, 6) else "left"),
            )
        # Status colour
        fill, font = _REPORT_STYLE.get(res.status, ("FFFFFF", "000000"))
        sc_cell = ws.cell(i, 3)
        sc_cell.fill = PatternFill("solid", fgColor=fill)
        sc_cell.font = Font(color=font, bold=True)
        sc_cell.alignment = Alignment(vertical="top", horizontal="center")
        # Clickable source link
        if res.url:
            link = ws.cell(i, 7)
            link.hyperlink = res.url
            link.font = Font(color="0563C1", underline="single")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(1, len(rows) + 1)}"


# --------------------------------------------------------------------------- #
# CLI entry point
# --------------------------------------------------------------------------- #
if __name__ == "__main__":
    import argparse

    here = os.path.dirname(os.path.abspath(__file__))
    default_root = os.path.normpath(os.path.join(here, "..", "Products"))
    default_csv = os.path.normpath(os.path.join(here, "..", "Item List - 260327 (Items with ID).csv"))

    ap = argparse.ArgumentParser(description="Pulse Rx description scraper (CLI)")
    ap.add_argument("--root", default=default_root)
    ap.add_argument("--csv", default=default_csv)
    ap.add_argument("--offset", type=int, default=0)
    ap.add_argument("--limit", type=int, default=10)
    ap.add_argument("--write", action="store_true", help="write results to xlsx")
    ap.add_argument("--order", default="dwatson,healthwire")
    ap.add_argument("--min-score", type=float, default=80.0)
    ap.add_argument("--delay", type=float, default=0.8)
    ap.add_argument("--threads", type=int, default=4, help="concurrent products (1 = sequential)")
    ap.add_argument("--report", default="", help="write a styled xlsx report to this path")
    args = ap.parse_args()

    cfg = ScrapeConfig(
        products_root=args.root, csv_path=args.csv,
        source_order=tuple(args.order.split(",")),
        min_score=args.min_score, request_delay=args.delay,
        max_workers=args.threads,
    )
    engine = ScraperEngine(cfg)
    folders = engine.select_folders(args.offset, args.limit)
    print(f"Processing {len(folders)} products (offset={args.offset})")

    def cb(res: MatchResult):
        print(f"[{res.status:11}] {res.folder_id} | {res.source or '-':10} "
              f"score={res.score:5} | {res.matched_name[:45]:45} | {res.csv_name[:40]}")
        if args.write and res.status == STATUS_FOUND and res.description:
            write_description(cfg.products_root, res.folder_id, res.description, cfg.header)

    results = engine.run_batch(folders, on_result=cb)
    found = sum(1 for r in results if r.status == STATUS_FOUND)
    print(f"\nDone: {found}/{len(results)} found.")

    if args.report:
        from datetime import datetime
        export_report(results, args.report, meta={
            "generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "products_root": cfg.products_root,
            "source_order": ", ".join(cfg.source_order),
            "threads": cfg.max_workers,
        })
        print(f"Report written: {args.report}")
